"""WhatsApp Web message sender and smart data extractor.

Assignment requirements covered:
- Playwright drives a real WhatsApp Web browser.
- Contacts are loaded from contacts.xlsx (Name, Phone, Message).
- Messages support the {name} placeholder.
- Random 2-5 second waits are used between important actions.
- Every successful send receives a screenshot.
- The last three incoming messages are extracted.
- Dated JSON and Excel reports are written after every run.

Install:
    python -m pip install playwright openpyxl
    python -m playwright install chromium

Run:
    python playwright_assign.py

Only message people who have explicitly agreed to hear from you. Automated
or high-volume messaging can violate WhatsApp rules and may restrict an account.
"""

from __future__ import annotations

import json
import os
import random
import re
import sys
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from playwright.sync_api import (
        BrowserContext,
        Locator,
        Page,
        TimeoutError as PlaywrightTimeoutError,
        sync_playwright,
    )
except ImportError:
    print("Missing dependencies. Run these commands in the active environment:")
    print("  python -m pip install playwright openpyxl")
    print("  python -m playwright install chromium")
    raise SystemExit(1)


# ----------------------------- Configuration -----------------------------

BASE_DIR = Path(__file__).resolve().parent
CONTACTS_FILE = BASE_DIR / "contacts.xlsx"
OUTPUT_DIR = BASE_DIR / "whatsapp_outputs"
SCREENSHOT_DIR = OUTPUT_DIR / "screenshots"
PROFILE_DIR = BASE_DIR / ".whatsapp_playwright_profile"
WHATSAPP_URL = "https://web.whatsapp.com/"

DEFAULT_MESSAGE = "Hello {name}, this is your daily update."
MIN_ACTION_DELAY_SECONDS = 2.0
MAX_ACTION_DELAY_SECONDS = 5.0
LOGIN_TIMEOUT_MS = 180_000
ELEMENT_TIMEOUT_MS = 20_000
SEND_CONFIRMATION_TIMEOUT_MS = 20_000
HEADLESS = False

# Set WHATSAPP_DRY_RUN=1 to test contact navigation without sending messages.
DRY_RUN = os.environ.get("WHATSAPP_DRY_RUN", "0").strip() == "1"
# Disabled by default so multiple contacts reuse the same loaded WhatsApp tab.
ALLOW_DIRECT_NUMBER_FALLBACK = (
    os.environ.get("WHATSAPP_DIRECT_NUMBER_FALLBACK", "0").strip() == "1"
)


SEARCH_BOX_SELECTORS = (
    'input[role="textbox"][data-tab="3"]',
    'input[aria-label="Search or start a new chat"]',
    'input[placeholder="Search or start a new chat"]',
    'div[contenteditable="true"][data-tab="3"]',
    'div[contenteditable="true"][aria-label*="Search"]',
    'div[contenteditable="true"][role="textbox"][aria-placeholder*="Search"]',
)

MESSAGE_BOX_SELECTORS = (
    'footer div[contenteditable="true"][role="textbox"]',
    'footer div[contenteditable="true"][data-tab="10"]',
    'footer div[contenteditable="true"]',
)


# ------------------------------- Data model -------------------------------

@dataclass
class Contact:
    row_number: int
    name: str
    phone: str
    message_template: str


@dataclass
class ContactResult:
    row_number: int
    name: str
    phone: str
    personalized_message: str = ""
    status: str = "pending"
    sent_at: str = ""
    screenshot: str = ""
    last_3_received_messages: list[str] = field(default_factory=list)
    error: str = ""


# ------------------------------- Utilities --------------------------------

def timestamp() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def dated_name(prefix: str, suffix: str) -> str:
    return f"{prefix}_{datetime.now():%Y-%m-%d}{suffix}"


def sanitize_filename(value: str, fallback: str = "contact") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return (cleaned or fallback)[:80]


def normalize_phone(value: Any) -> str:
    raw = str(value or "").strip()
    digits = re.sub(r"\D", "", raw)
    return f"+{digits}" if digits else ""


def validate_phone(phone: str) -> str:
    if not re.fullmatch(r"\+\d{7,15}", phone):
        return "Phone must contain a country code and 7-15 digits."
    return ""


def personalize(template: str, name: str) -> str:
    source = template.strip() or DEFAULT_MESSAGE
    return source.replace("{name}", name)


def human_delay(page: Page, minimum: float = MIN_ACTION_DELAY_SECONDS) -> None:
    maximum = max(minimum, MAX_ACTION_DELAY_SECONDS)
    delay_ms = int(random.uniform(minimum, maximum) * 1000)
    page.wait_for_timeout(delay_ms)


def wait_for_first_visible(
    page: Page,
    selectors: tuple[str, ...],
    timeout_ms: int = ELEMENT_TIMEOUT_MS,
) -> Locator:
    """Use wait_for_selector explicitly and return the first visible locator."""
    per_selector_timeout = max(1_500, timeout_ms // len(selectors))
    errors: list[str] = []
    for selector in selectors:
        try:
            page.wait_for_selector(
                selector, state="visible", timeout=per_selector_timeout
            )
            return page.locator(selector).first
        except PlaywrightTimeoutError:
            errors.append(selector)
    raise RuntimeError("No visible element matched: " + " | ".join(errors))


def create_contacts_template() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"
    sheet.append(["Name", "Phone", "Message"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 55

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Field", "Requirement"])
    instructions.append(["Name", "Contact name used for {name} personalization"])
    instructions.append(["Phone", "Country code required, for example +6591234567"])
    instructions.append(["Message", "Optional; blank uses the built-in default template"])
    instructions.append(
        ["Consent", "Only add contacts who agreed to receive these messages"]
    )
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 72
    workbook.save(CONTACTS_FILE)


def load_contacts() -> list[Contact]:
    workbook = load_workbook(CONTACTS_FILE, data_only=True)
    if "Contacts" not in workbook.sheetnames:
        raise ValueError("contacts.xlsx must contain a sheet named 'Contacts'.")
    sheet = workbook["Contacts"]
    headers = {
        str(cell.value or "").strip().lower(): index
        for index, cell in enumerate(sheet[1], start=1)
    }
    missing = [name for name in ("name", "phone", "message") if name not in headers]
    if missing:
        raise ValueError("contacts.xlsx is missing columns: " + ", ".join(missing))

    contacts: list[Contact] = []
    for row_number in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row_number, headers["name"]).value or "").strip()
        phone = normalize_phone(sheet.cell(row_number, headers["phone"]).value)
        message = str(
            sheet.cell(row_number, headers["message"]).value or ""
        ).strip()
        if not name and not phone and not message:
            continue
        contacts.append(Contact(row_number, name, phone, message))
    return contacts


# ---------------------------- WhatsApp actions ----------------------------

def wait_for_login(page: Page) -> None:
    page.goto(WHATSAPP_URL, wait_until="domcontentloaded", timeout=60_000)
    try:
        page.wait_for_selector("#pane-side", state="visible", timeout=15_000)
        print("WhatsApp Web session restored from the persistent profile.")
        return
    except PlaywrightTimeoutError:
        print("Scan the WhatsApp Web QR code with your phone.")
        print("Waiting up to 3 minutes for login...")

    page.wait_for_selector("#pane-side", state="visible", timeout=LOGIN_TIMEOUT_MS)
    page.wait_for_timeout(2_000)
    print("WhatsApp Web login confirmed.")


def clear_search(page: Page) -> None:
    try:
        search_box = wait_for_first_visible(page, SEARCH_BOX_SELECTORS, 8_000)
        search_box.click()
        page.keyboard.press("Control+A")
        page.keyboard.press("Backspace")
        page.wait_for_timeout(500)
    except Exception:
        pass


def click_matching_search_result(page: Page, query: str) -> bool:
    pane = page.locator("#pane-side")
    query_digits = re.sub(r"\D", "", query)
    query_lower = query.lower()

    row_selectors = (
        '#pane-side [role="listitem"]',
        '#pane-side [role="row"]',
        '#pane-side div[tabindex="-1"]',
    )
    for selector in row_selectors:
        rows = page.locator(selector)
        count = min(rows.count(), 30)
        for index in range(count):
            row = rows.nth(index)
            try:
                text = row.inner_text(timeout=1_500).strip()
            except Exception:
                continue
            text_digits = re.sub(r"\D", "", text)
            matched = query_lower in text.lower()
            if query_digits:
                matched = matched or query_digits[-8:] in text_digits
            if matched:
                row.click(timeout=8_000)
                return True

    # Current WhatsApp phone searches may display only the saved contact name,
    # not the phone digits. The filtered list starts with a "Chats" heading,
    # followed by the actual matching contact rows.
    filtered_rows = page.locator('#pane-side [role="row"]')
    for index in range(min(filtered_rows.count(), 30)):
        row = filtered_rows.nth(index)
        try:
            if not row.is_visible():
                continue
            text = row.inner_text(timeout=1_500).strip()
        except Exception:
            continue
        normalized = re.sub(r"\s+", " ", text).strip().lower()
        if not normalized or normalized in {"chats", "contacts", "messages"}:
            continue
        if "no chats, contacts or messages found" in normalized:
            return False
        row.click(timeout=8_000)
        return True

    try:
        text_match = pane.get_by_text(query, exact=False).first
        text_match.wait_for(state="visible", timeout=4_000)
        text_match.click()
        return True
    except Exception:
        return False


def search_and_open_contact(page: Page, contact: Contact) -> bool:
    queries = [contact.phone] if contact.phone else [contact.name]
    for query in queries:
        clear_search(page)
        search_box = wait_for_first_visible(page, SEARCH_BOX_SELECTORS)
        search_box.click()
        search_box.fill(query)
        human_delay(page)
        if click_matching_search_result(page, query):
            try:
                wait_for_first_visible(page, MESSAGE_BOX_SELECTORS, 12_000)
                return True
            except Exception:
                continue

    # Optional support for unsaved numbers. It reloads WhatsApp, so it remains
    # disabled by default to keep multi-contact runs in one loaded web session.
    if contact.phone and ALLOW_DIRECT_NUMBER_FALLBACK:
        digits = re.sub(r"\D", "", contact.phone)
        page.goto(
            f"https://web.whatsapp.com/send?phone={digits}",
            wait_until="domcontentloaded",
            timeout=45_000,
        )
        try:
            wait_for_first_visible(page, MESSAGE_BOX_SELECTORS, 20_000)
            return True
        except Exception:
            return False
    return False


def send_message(page: Page, message: str) -> None:
    message_box = wait_for_first_visible(page, MESSAGE_BOX_SELECTORS)
    outgoing = page.locator(".message-out")
    before_count = outgoing.count()

    message_box.click()
    message_box.fill(message)
    page.wait_for_timeout(random.randint(800, 1_500))
    if DRY_RUN:
        print("DRY RUN: message prepared but not sent.")
        message_box.fill("")
        return

    message_box.press("Enter")
    page.wait_for_function(
        "before => document.querySelectorAll('.message-out').length > before",
        arg=before_count,
        timeout=SEND_CONFIRMATION_TIMEOUT_MS,
    )
    page.wait_for_timeout(1_500)


def extract_last_three_incoming(page: Page) -> list[str]:
    page.wait_for_selector("#main", state="visible", timeout=10_000)
    incoming = page.locator("#main .message-in")
    count = incoming.count()
    messages: list[str] = []
    for index in range(max(0, count - 3), count):
        bubble = incoming.nth(index)
        text_locator = bubble.locator("span.selectable-text").last
        try:
            text = text_locator.inner_text(timeout=3_000).strip()
        except Exception:
            try:
                text = bubble.inner_text(timeout=3_000).strip()
            except Exception:
                text = ""
        if text:
            messages.append(text)
    return messages[-3:]


def screenshot_sent_chat(page: Page, contact: Contact) -> Path:
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    identity = contact.name or contact.phone or f"row_{contact.row_number}"
    filename = (
        f"{sanitize_filename(identity)}_{datetime.now():%Y-%m-%d_%H%M%S}.png"
    )
    path = SCREENSHOT_DIR / filename
    page.screenshot(path=str(path), full_page=False)
    return path


def process_contact(page: Page, contact: Contact) -> ContactResult:
    result = ContactResult(contact.row_number, contact.name, contact.phone)
    phone_error = validate_phone(contact.phone)
    if not contact.name:
        result.status = "invalid"
        result.error = "Name is required."
        return result
    if phone_error:
        result.status = "invalid"
        result.error = phone_error
        return result

    result.personalized_message = personalize(contact.message_template, contact.name)
    print(f"\nRow {contact.row_number}: {contact.name} ({contact.phone})")
    try:
        if not search_and_open_contact(page, contact):
            result.status = "contact_not_found"
            result.error = "No WhatsApp chat or valid number was found."
            return result

        human_delay(page)
        send_message(page, result.personalized_message)
        if DRY_RUN:
            result.status = "dry_run"
            return result

        result.sent_at = timestamp()
        result.status = "sent"
        human_delay(page)
        screenshot_path = screenshot_sent_chat(page, contact)
        result.screenshot = str(screenshot_path.relative_to(BASE_DIR))

        try:
            result.last_3_received_messages = extract_last_three_incoming(page)
        except Exception as extraction_error:
            result.error = f"Message sent; extraction failed: {extraction_error}"
        return result
    except PlaywrightTimeoutError as error:
        result.status = "failed"
        result.error = f"Timed out waiting for WhatsApp Web: {error}"
        return result
    except Exception as error:
        result.status = "failed"
        result.error = str(error)
        return result
    finally:
        try:
            clear_search(page)
        except Exception:
            pass


# ------------------------------- Reporting --------------------------------

def write_json_report(results: list[ContactResult], started_at: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / dated_name("whatsapp_report", ".json")
    payload = {
        "run": {
            "started_at": started_at,
            "completed_at": timestamp(),
            "dry_run": DRY_RUN,
            "total_contacts": len(results),
            "sent": sum(item.status == "sent" for item in results),
            "failed": sum(
                item.status in {"failed", "contact_not_found", "invalid"}
                for item in results
            ),
        },
        "contacts": [asdict(item) for item in results],
    }
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def write_excel_report(results: list[ContactResult]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / dated_name("whatsapp_report", ".xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WhatsApp Report"
    headers = [
        "Row",
        "Name",
        "Phone",
        "Personalized Message",
        "Status",
        "Sent At",
        "Screenshot",
        "Received Message 1",
        "Received Message 2",
        "Received Message 3",
        "Error",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for result in results:
        incoming = result.last_3_received_messages + ["", "", ""]
        sheet.append(
            [
                result.row_number,
                result.name,
                result.phone,
                result.personalized_message,
                result.status,
                result.sent_at,
                result.screenshot,
                incoming[0],
                incoming[1],
                incoming[2],
                result.error,
            ]
        )

    widths = [8, 22, 20, 45, 20, 26, 45, 38, 38, 38, 55]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    return path


def write_reports(results: list[ContactResult], started_at: str) -> None:
    json_path = write_json_report(results, started_at)
    excel_path = write_excel_report(results)
    print(f"JSON report:  {json_path}")
    print(f"Excel report: {excel_path}")


# --------------------------------- Main ------------------------------------

def run_browser(contacts: list[Contact], results: list[ContactResult]) -> None:
    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as playwright:
        context: BrowserContext = playwright.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=HEADLESS,
            viewport=None,
            args=["--start-maximized"],
        )
        context.set_default_timeout(ELEMENT_TIMEOUT_MS)
        page = context.pages[0] if context.pages else context.new_page()
        try:
            wait_for_login(page)
            for position, contact in enumerate(contacts, start=1):
                print(f"Processing {position}/{len(contacts)}...")
                results.append(process_contact(page, contact))
                if position < len(contacts):
                    human_delay(page)
        finally:
            context.close()


def main() -> int:
    print("WhatsApp Message Sender + Smart Data Extractor")
    print("Use responsibly and only with contacts who provided consent.")
    if DRY_RUN:
        print("DRY RUN is enabled: no messages will be sent.")

    if not CONTACTS_FILE.exists():
        create_contacts_template()
        print(f"Created contacts template: {CONTACTS_FILE}")
        print("Add consenting contacts, save the workbook, and run the script again.")
        return 0

    try:
        contacts = load_contacts()
    except Exception as error:
        print(f"Could not read contacts.xlsx: {error}")
        return 1
    if not contacts:
        print("contacts.xlsx contains no contact rows.")
        return 1

    started_at = timestamp()
    results: list[ContactResult] = []
    exit_code = 0
    try:
        run_browser(contacts, results)
    except PlaywrightTimeoutError as error:
        print(f"WhatsApp login or browser operation timed out: {error}")
        exit_code = 1
    except KeyboardInterrupt:
        print("Run interrupted by the user. Writing partial reports...")
        exit_code = 2
    except Exception as error:
        print(f"Browser automation stopped: {error}")
        exit_code = 1
    finally:
        processed_rows = {result.row_number for result in results}
        for contact in contacts:
            if contact.row_number not in processed_rows:
                results.append(
                    ContactResult(
                        contact.row_number,
                        contact.name,
                        contact.phone,
                        personalized_message=personalize(
                            contact.message_template, contact.name
                        ),
                        status="not_processed",
                        error="The run stopped before this contact was processed.",
                    )
                )
        results.sort(key=lambda item: item.row_number)
        write_reports(results, started_at)

    sent = sum(result.status == "sent" for result in results)
    print(f"Completed: {sent}/{len(results)} messages sent.")
    print(f"Persistent login profile: {PROFILE_DIR}")
    print("Do not upload the persistent login profile to GitHub.")
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
